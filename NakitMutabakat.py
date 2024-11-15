import sys
import os
import re
import pyodbc
import pandas as pd
from datetime import datetime, timedelta
#from openpyxl import load_workbook, Workbook
import openpyxl
import shutil

#Ekranı temizle
clear = lambda: os.system('cls')
clear()


#====  A Y A R L A R  ==========================================================================
sql_sunucu        = '*****'
sql_veritabani    = 'PaygateMaestro'
sql_kullanici     = '***r'
sql_sifre         = '**********'

# Refrans dosyasında bulunan fakat Swift dosyasına yazılması istenmeyen Referans'lar
swift_excel_ref_filitresi = ['KURUM','MÜŞTERİ','takas farlı','takas farkı','TAKAS FARKI']

# Manuel Tarih Ayarı 
# Manel tarih girmek için = 'gg.aa.yyyy' formatında tarih giriniz
# Manel tarihi silmek için değişkenlere ='' atayın
# Manuel tarih girerken oluşturulma tarihi bir gün sonraya kayan kayıtlarında
# gelmesi için +1 gün eklemeyi unutma!
manuel_tarih_baslangici = '' #'20.09.2024'
manuel_tarih_bitisi     = '' #'28.09.2024'

#====  A Y A R L A R  ==========================================================================



# Define transaction type codes
transaction_types = {
    'S': 'SWIFT transfer',
    'N': 'Non-SWIFT transfer',
    'F': 'First advice'
}

# Define identification codes for Transaction Types 'N' or 'F'
identification_codes = {
    'BNK': 'Securities Related Item - Bank Fees',
    'BOE': 'Bill of Exchange',
    'BRF': 'Brokerage Fee',
    'CAR': 'Securities Related Item - Corporate Actions Related',
    'CAS': 'Securities Related Item - Cash in Lieu',
    'CHG': 'Charges and Other Expenses',
    'CHK': 'Cheques',
    'CLR': 'Cash Letters/Cheques Remittance',
    'CMI': 'Cash Management Item - No Detail',
    'CMN': 'Cash Management Item - Notional Pooling',
    'CMP': 'Compensation Claims',
    'CMS': 'Cash Management Item - Sweeping',
    'CMT': 'Cash Management Item - Topping',
    'CMZ': 'Cash Management Item - Zero Balancing',
    'COL': 'Collections',
    'COM': 'Commission',
    'CPN': 'Securities Related Item - Coupon Payments',
    'DCR': 'Documentary Credit',
    'DDT': 'Direct Debit Item',
    'DIS': 'Securities Related Item - Gains Disbursement',
    'DIV': 'Securities Related Item - Dividends',
    'EQA': 'Equivalent Amount',
    'EXT': 'Securities Related Item - External Transfer for Own Account',
    'FEX': 'Foreign Exchange',
    'INT': 'Interest Related Amount',
    'LBX': 'Lock Box',
    'LDP': 'Loan Deposit',
    'MAR': 'Securities Related Item - Margin Payments/Receipts',
    'MAT': 'Securities Related Item - Maturity',
    'MGT': 'Securities Related Item - Management Fees',
    'MSC': 'Miscellaneous',
    'NWI': 'Securities Related Item - New Issues Distribution',
    'ODC': 'Overdraft Charge',
    'OPT': 'Securities Related Item - Options',
    'PCH': 'Securities Related Item - Purchase',
    'POP': 'Securities Related Item - Pair-off Proceeds',
    'PRN': 'Securities Related Item - Principal Pay-down/Pay-up',
    'REC': 'Securities Related Item - Tax Reclaim',
    'RED': 'Securities Related Item - Redemption/Withdrawal',
    'RIG': 'Securities Related Item - Rights',
    'RTI': 'Returned Item',
    'SAL': 'Securities Related Item - Sale',
    'SEC': 'Securities',
    'SLE': 'Securities Related Item - Securities Lending Related',
    'STO': 'Standing Order',
    'STP': 'Securities Related Item - Stamp Duty',
    'SUB': 'Securities Related Item - Subscription',
    'SWP': 'Securities Related Item - SWAP Payment',
    'TAX': 'Securities Related Item - Withholding Tax Payment',
    'TCK': 'Travellers Cheques',
    'TCM': 'Securities Related Item - Tripartite Collateral Management',
    'TRA': 'Securities Related Item - Internal Transfer for Own Account',
    'TRF': 'Transfer',
    'TRN': 'Securities Related Item - Transaction Fee',
    'UWC': 'Securities Related Item - Underwriting Commission',
    'VDA': 'Value Date Adjustment',
    'WAR': 'Securities Related Item - Warrant'
}



#===============================================================================================
# Veritabanı bağlantı fonksiyonu
def veritabani_baglan(sunucu, veritabani, kullanici_adi, sifre):
    try:
        driver_name = ''
        driver_names = [x for x in pyodbc.drivers() if x.endswith('for SQL Server')]
        if driver_names:
            driver_name= '{' + driver_names[0] + '}'

        if driver_name:
            connStr = f"DRIVER={driver_name};TrustServerCertificate=yes;SERVER={sunucu};DATABASE={veritabani};UID={kullanici_adi};PWD={sifre}"
            conn = pyodbc.connect(connStr)
            return conn
        else:
            print("Uygun ODBC sürücüsü bulunamadı! SQL ODBC Sürücüsü oldundan emin olun!")
            sys.exit(1)

    except pyodbc.Error as e:
        print(f"DB bağlantı hatası: {e}")
        sys.exit(1)
#===============================================================================================


#===============================================================================================
# Swift 62F Mesajını parse et
def parse_62F_field(field,gunsonu_tarihi):
    # SWIFT 62F pattern
    pattern = r":62F:([CD])(\d{6})([A-Z]{3})([\d,]+)"
    
    # Match the pattern
    match = re.match(pattern, field)
    
    if match:
        balance_type = match.group(1)  # 'C' for Credit, 'D' for Debit
        date_str     = match.group(2)  # Date in YYMMDD format
        currency     = match.group(3)  # Currency code (e.g. USD, EUR)
        amount_str   = match.group(4)  # Amount as a string
        
        # Convert date to a readable format
        date_obj = datetime.strptime(date_str, "%y%m%d").date()
        
        # Convert amount to a float (handle comma for decimal places)
        amount = float(amount_str.replace(",", "."))

        if(gunsonu_tarihi.date() == date_obj):
            # Create a dictionary with parsed information
            result = {
                "balance_type": balance_type,
                "date": date_obj,
                "currency": currency,
                "amount": amount
            }
            return result
        else:
            return None
    else:
        raise ValueError("Invalid 62F field format")
#===============================================================================================


#===============================================================================================
# Swift 61 Mesajını parse et
def parse_61_field(field,curr):
    # SWIFT 61 pattern
    pattern = r":61:(\d{6})(\d{4})?([CD])([\d,]+)([A-Z]{4})(.*)"
    
    # Match the pattern
    match = re.match(pattern, field)
    
    if match:
        # Valör tarihi (işlemin geçerli olduğu tarih)
        value_date_str = match.group(1)
        # Muhasebe tarihi (opsiyonel, genellikle belirtilir)
        entry_date_str = match.group(2)
        # İşlem türü (C for Credit, D for Debit)
        debit_credit = match.group(3)
        # İşlem tutarı
        amount_str = match.group(4)
        # İşlem detayı (örneğin, NTRF gibi işlem türünü gösterir)
        parsed_transaction_type = parse_transaction_type(match.group(5))
        # İşlem detayını parse ederek transaction_type ve identification_code'u getir
        transaction_type = parsed_transaction_type["transaction_type"]
        identification_code = parsed_transaction_type["identification_code"]
        # İşlem referansı ve ek bilgiler
        reference_info = match.group(6)

        reference_end = reference_info.find("//", 0)
        reference = reference_info[0:reference_end].split(',')[0].rstrip() 

        # Valör tarihini dönüştür
        value_date = datetime.strptime(value_date_str, "%y%m%d").date()

        # Muhasebe tarihi varsa dönüştür
        entry_date = None
        if entry_date_str:
            entry_date = datetime.strptime(entry_date_str, "%m%d").date()
            # Muhasebe tarihine yıl ekleme
            entry_date = entry_date.replace(year=value_date.year)

        # Tutarı float'a dönüştür
        amount = float(amount_str.replace(",", "."))
        
        # Sonuç olarak bir dictionary döndür
        result = {
            "REFERENCE"   : reference,
            "CURR"        : curr,
            "CREDIT/DEBIT": debit_credit,
            "AMOUNT"      : amount,
            "TRANSTYPE"   : transaction_type,
            "IDCODE"      : identification_code,
            "VALDATE"     : value_date,
            "ENTDATE"     : entry_date
        }
        
        return result
    else:
        raise ValueError("Invalid 61 field format")
#===============================================================================================



#===============================================================================================
def parse_transaction_type(transaction_type_raw):
    # Extract transaction type (first character of the input)
    transaction_type_code = transaction_type_raw[0]
    transaction_type = transaction_types.get(transaction_type_code)

    # Extract identification code if transaction type is 'N' or 'F'
    identification_code = None
    if transaction_type in ['Non-SWIFT transfer', 'First advice']:
        identification_code_code = transaction_type_raw[1:4]
        identification_code = identification_codes.get(identification_code_code)

    return {
        'transaction_type': transaction_type,
        'identification_code': identification_code
    }
#===============================================================================================


#===============================================================================================
def MutabakatYap(ana_dizin,dosya_on_eki,SenderBIC,manuel_tarih_baslangici,manuel_tarih_bitisi,parcali_islemleri_birlestir):

    # Manuel tarih girildiyse kontrol et!
    if(len(manuel_tarih_baslangici) > 0 or len(manuel_tarih_bitisi) > 0):
        try:
            tarih_baslangici = (datetime.strptime(manuel_tarih_baslangici, '%d.%m.%Y')).replace(hour=0, minute=0, second=0, microsecond=0)   
        except:
            clear()
            print(f"Hatalı başlangıç tarih tanımı! {manuel_tarih_baslangici} tarihe çevrilemedi!")
            exit()

        try:
            tarih_bitisi = (datetime.strptime(manuel_tarih_bitisi, '%d.%m.%Y')).replace(hour=6, minute=0, second=0, microsecond=0)

        except:
            clear()
            print(f"Hatalı bitiş tarih tanımı! {manuel_tarih_bitisi} tarihe çevrilemedi!")
            exit()

        # UNUTMAAA!!!!!!!!!!!!   Manuel tarih tanımı yapıldığı durumda cikartilacak_gun_sayisi'na default bir değer atıyoruz!!!!!!1
        #cikartilacak_gun_sayisi = 1

        print("Manuel tarih tanımı yapılmış")
        print("")
        print(f"Sorgu {manuel_tarih_baslangici} - {manuel_tarih_bitisi} tarihi arasında çalışacaktır!")
        print("")
    
    else:
        # Mevcut tarihi al
        simdi = datetime.now()
        if simdi.weekday() == 0:
            cikartilacak_gun_sayisi = 3
        else:
            cikartilacak_gun_sayisi = 1

        # Bir gün öncenin tarihini getir
        bir_gun_once = simdi - timedelta(days=cikartilacak_gun_sayisi)
        # SQL sorgusu başlangıç saatini bir gün öncenin 00:00:00 saati olarak al
        tarih_baslangici  = bir_gun_once.replace(hour=0, minute=0, second=0, microsecond=0)
        tarih_bitisi      = simdi.replace(hour=6, minute=0, second=0, microsecond=0)

    sql_baglantisi = veritabani_baglan(sql_sunucu, sql_veritabani, sql_kullanici, sql_sifre)  
    sorgu =  "SELECT [MM].[MTID],[MM].[Reference],[MM].[SenderBIC],[MM].[ReceiverBIC],[MM].[Currency],[MM].[Amount] "
    sorgu += ",[MM].[ValueDate],[MM].[RecordDate],[MM].[CreationDate],[MR].[Header],[MR].[Body],[MR].[Trailer] "
    sorgu += "FROM [dbo].[PGM_MessagesMaster] MM "
    sorgu += "INNER JOIN [dbo].[PGM_MessagesRaw] MR ON [MM].[MTID] = [MR].[MTID] "
    sorgu += "WHERE [MM].[MessageType] = 950 "
    sorgu += f"AND [MM].[RecordDate] BETWEEN '{tarih_baslangici}' AND '{tarih_bitisi}' "
    sorgu += f"AND [MM].[SenderBIC] = '{SenderBIC}'"
    # sorgu += f"AND [MM].[ValueDate] = '2024-10-30'"
    # sorgu += f"AND [MM].[ValueDate] = BETWEEN '{tarih_baslangici}' AND '{tarih_bitisi}'"
    df_swift_data = pd.read_sql(sorgu, sql_baglantisi) 

    if len(df_swift_data) ==0  :
        clear()
        print(f"DİKKAT! Swift '{tarih_baslangici} - {tarih_bitisi} arasında {SenderBIC} için hiç kayıt döndürmedi! ")
        print('====================================================================================')
        print(f"Sorgu: {sorgu}")
        return 

    lines_with_61  = []
    lines_with_62f = []

    #Sorgudan dönen her bir kaydı dön
    for index, row in df_swift_data.iterrows():
        # Kaydın Body 'sini her bir yeni satırı bir line olacak şekilde böl
        lines = row['Body'].split('\n')
        # Body içindeki her bir yeni satırı dön
        for line in lines:
            # Satır :61: tag'ını içeriyorsa parse et ve lines_with_61 dictionary array'ine ekle
            if ":61:" in line:
                lines_with_61.append(parse_61_field(line,row['Currency']))
            # Satır :62F: tag'ını içeriyorsa parse et ve lines_with_62f dictionary array'ine ekle
            if ":62F:" in line:
                parsed_62F = parse_62F_field(line,tarih_bitisi - timedelta(days=cikartilacak_gun_sayisi) )
                if(parsed_62F is not None):
                    lines_with_62f.append(parsed_62F)

    # lines_with_61 dictionary array'ini dataframe'e çevir
    df_61  = pd.DataFrame.from_dict(lines_with_61) 


    # Paraçlı işlemleri aynı referans altına birleştir
    if (parcali_islemleri_birlestir == True):
        df_61 = df_61.groupby(['REFERENCE', 'CURR', 'CREDIT/DEBIT','TRANSTYPE','IDCODE','VALDATE','ENTDATE'], as_index=False).agg({'AMOUNT': 'sum'})


    # lines_with_62f dictionary array'ini dataframe'e çevir
    df_62f = pd.DataFrame.from_dict(lines_with_62f)

    if(len(df_62f) < 1):
        print(f"Belirtilen tarih için hiç 62 kaydı bulunamadı!")
    else:
        # currency sutununu index'e çevir
        df_62f.set_index("currency",inplace = True)

    #Refrans Excel dosyasını aç     
    bugun = datetime.today()
    tarih_string = bugun.strftime("%d%m%Y")

    #Referans dosya adını belirle
    referans_dosya_adi = f"{ana_dizin}/{dosya_on_eki} {tarih_string}.xlsx"
    yedek_referans_dosya_adi = referans_dosya_adi.replace('.xlsx','_yedek.xlsx')

    # Referans dosyasını yedekle
    if os.path.exists(yedek_referans_dosya_adi):
        os.remove(yedek_referans_dosya_adi)
    
    shutil.copy(referans_dosya_adi, yedek_referans_dosya_adi)
    print(f"Referans dosyası [{referans_dosya_adi}] [{yedek_referans_dosya_adi}] dosyası olarak yedeklendi")

    # Referans dosyasını aç
    if os.path.exists(referans_dosya_adi):
        referans_workbook = openpyxl.load_workbook(referans_dosya_adi)
    else:
        print(f"Referans dosyası [{referans_dosya_adi}] bulunamadı!")
        exit(0)
    
    # Swift dosyasının adını belirle
    swift_dosya_adi = f"{ana_dizin}/{dosya_on_eki}_SWIFT_{tarih_string}.xlsx"

    #Eğer Swift dosyası mevcut ise sil
    if os.path.isfile(swift_dosya_adi):
        os.remove(swift_dosya_adi)

    #Her bir para birimine göre dön 
    for currency in df_61['CURR'].unique():
        currency_upper = currency.upper()
        
        # Refreans excel dosyasının geçerli (currency_upper ile belirlenene) sayfasını datafreame'e yükle
        df_referans = pd.read_excel(referans_dosya_adi,sheet_name=currency_upper)
        # Referans dosyasındaki satır numarasını index'den al.
        # Daha sonra dosyayı güncellerken hangi satırın güncelleneceğini bu satır -
        # sutun üzerinden belirleyeceğiz
        df_referans['SatNo'] = df_referans.index + 2
        # Gereksiz sutunları sil, sadece REFERENCE, DEBIT, CREDIT ve SatNo sutunları kalsın
        df_referans = df_referans[['REFERENCE','DEBIT','CREDIT','SatNo']]

        # REFERENCE sutunu boş olan satırları sil. Bu satırlar eşleştirmede kullanılmayacaktır
        df_referans = df_referans.dropna(subset=['REFERENCE'])

        # Referasn excel'indeki veri ile Swift verinin ortak alanı olan REFERENCE sutununu 
        # object'den string'e çevir ki merge oprasyonu düzgün çalışsın
        df_referans['REFERENCE'] = df_referans['REFERENCE'].astype('string')

        #  DEBIT ve CREDIT sutunlarını numeriğe çevir (NaN'ları sil)
        df_referans['DEBIT']  = pd.to_numeric(df_referans['DEBIT'], errors='coerce')
        df_referans['CREDIT'] = pd.to_numeric(df_referans['CREDIT'], errors='coerce')

        # df_61 ve df_referans datasını sadece df_referans'da olan kayıtları 
        # içerecek şekilde REFERENCE sutununu baz alarak birleştir
        df_ref_merged = pd.DataFrame
        df_ref_merged = pd.merge(df_referans, df_61, on='REFERENCE', how='left')

        # df_ref_merged için DEBIT, CREDIT ve AMOUNT sutunlarını noktandan 
        # sonra 2 hane olacak şekilde yuvarla
        df_ref_merged['DEBIT']  = df_ref_merged['DEBIT'].round(2)
        df_ref_merged['CREDIT'] = df_ref_merged['CREDIT'].round(2)
        df_ref_merged['AMOUNT'] = df_ref_merged['AMOUNT'].round(2)

        # Referans dataframe'i için DEBIT_FARK ve CREDIT_FARK'ı hesapla
        df_ref_merged['DEBIT_FARK'] = df_ref_merged.apply(
            lambda row: row['DEBIT'] - row['AMOUNT'] if row['CREDIT/DEBIT'] == 'D' else None, axis=1)
        df_ref_merged['CREDIT_FARK'] = df_ref_merged.apply(
            lambda row: row['CREDIT'] - row['AMOUNT'] if row['CREDIT/DEBIT'] == 'C' else None, axis=1)

        # Swift dosyalarını oluşturmak için gerekli dataframe'i oluştur
        # df_61 dataframe'ini sadece geçerli currency için filitrele 
        df_61_filtered=df_61.query(f"CURR == '{currency_upper}'")
        # df_referans ve df_61_filtered datafreame'lerini her iki tablodaki tüm kayıtları 
        # içerecek şekilde REFERENCE sutununu baz alarak birleştir
        df_swift_merged = pd.merge(df_referans ,df_61_filtered, on='REFERENCE', how='outer')

        df_swift_merged = df_swift_merged.dropna(subset=['AMOUNT'])

        # df_swift_merged dataframe'imde  swift_excel_ref_filitresi içinde bulunan satırları sil
        # df_swift_merged.drop(df_swift_merged.loc[df_swift_merged['REFERENCE'].isin(swift_excel_ref_filitresi)].index, inplace = True)
        # df_swift_merged için gereksiz SatNo ve CURR sutunlarını sil
        df_swift_merged = df_swift_merged.drop(columns=['SatNo','CURR'])

        # df_swift_merged için DEBIT, CREDIT ve AMOUNT sutunlarını noktandan 
        # sonra 2 hane olacak şekilde yuvarla
        df_swift_merged['DEBIT']  = df_swift_merged['DEBIT'].round(2)
        df_swift_merged['CREDIT'] = df_swift_merged['CREDIT'].round(2)
        df_swift_merged['AMOUNT'] = df_swift_merged['AMOUNT'].round(2)

        df_swift_merged.rename(columns={"AMOUNT" : "SWIFT_AMOUNT"},inplace=True)

        df_swift_merged['DEBIT'] = df_swift_merged['DEBIT'].fillna(0)
        df_swift_merged['CREDIT'] = df_swift_merged['CREDIT'].fillna(0)
        df_swift_merged['REF_AMOUNT'] = df_swift_merged['DEBIT'] + df_swift_merged['CREDIT']
        df_swift_merged['FARK'] = df_swift_merged['REF_AMOUNT'] - df_swift_merged['SWIFT_AMOUNT']
        
        # Excel'e yazdırıldığında daha doğru görünmesi için 
        # df_swift_merged için sutun sırasını değiştir
        #df_swift_merged = df_swift_merged[['REFERENCE','DEBIT','CREDIT','CREDIT/DEBIT','AMOUNT','DEBIT_FARK','CREDIT_FARK','TRANSTYPE','IDCODE']]
        df_swift_merged = df_swift_merged[['REFERENCE','CREDIT/DEBIT','REF_AMOUNT','SWIFT_AMOUNT','FARK','TRANSTYPE','IDCODE']]

        # df_swift_merged dataframe'ini excel'e yaz
        # Eğer dosya yoksa oluşturmak için ExcelWriter'ın modunu 'w' (üzerine yaz / oluştur) olarak belirle
        # Eğer dosya varsa ilgili sembolün sayfasını eklemek için ExcelWriter'ın modunu 'a' (ekle) olarak belirle
        yazma_modu = 'a' if os.path.isfile(swift_dosya_adi) else 'w'
        with pd.ExcelWriter(swift_dosya_adi, engine='openpyxl', mode=yazma_modu) as writer:
            df_swift_merged.to_excel(writer, sheet_name=currency_upper, index=False)
        
        # Referans excel'inin ilgili currency sayfasını seç
        referans_curr_sheet = referans_workbook[currency_upper]
        #referans_workbook.active = referans_workbook[currency_upper]

        # Referans excel'inde önceki eşleşme hücrelerini temizle
        for row in referans_curr_sheet['E2:G1000']:
            for cell in row:
                cell.value = None

        # Referans excel'indeki gün sonu bakiyesini yaz
        if(len(df_62f) > 0):
            df_gunsonu_fiyati = df_62f.query(f"currency == '{currency_upper}'")
            if(len(df_gunsonu_fiyati) > 0):
                referans_curr_sheet.cell(row=2, column=4).value = df_gunsonu_fiyati['amount'].values[0]
            else:
                referans_curr_sheet.cell(row=2, column=4).value = (tarih_bitisi.strftime('%d.%M.%Y tarihine ait gün sonu bakiyesi bulunamadı!'))
        else:
            referans_curr_sheet.cell(row=2, column=4).value = (tarih_bitisi.strftime('%d.%M.%Y tarihine ait gün sonu bakiyesi bulunamadı!'))

        # Referans dosyasını satır satır güncelle 
        for satir in range(len(df_ref_merged)):
            sat_no        = df_ref_merged.loc[satir, 'SatNo']
            merged_amount = df_ref_merged.loc[satir, 'AMOUNT']
            debit_fark    = df_ref_merged.loc[satir, 'DEBIT_FARK']
            credit_fark   = df_ref_merged.loc[satir, 'CREDIT_FARK']
            
            if(pd.isna(merged_amount)):
                referans_curr_sheet.cell(row=sat_no, column=5).value = 'N/A'
                referans_curr_sheet.cell(row=sat_no, column=6).value = 'N/A'
                referans_curr_sheet.cell(row=sat_no, column=7).value = 'N/A'
            else:
                if df_ref_merged.loc[satir, 'CREDIT/DEBIT'] == 'C':
                    # 6. sutun CREDIT
                    referans_curr_sheet.cell(row=sat_no, column=6).value = merged_amount
                else:
                    # 5. sutun DEBIT
                    referans_curr_sheet.cell(row=sat_no, column=5).value = merged_amount

                referans_curr_sheet.cell(row=sat_no, column=7).value = debit_fark
                referans_curr_sheet.cell(row=sat_no, column=8).value = credit_fark

    # Referans dosyasını yaz ve kapat
    referans_workbook.save(referans_dosya_adi)
    referans_workbook.close()
#==============================================================================================




#==============================================================================================

# Euroclear için Mutabakat yap
MutabakatYap(dosya_on_eki = "ECL", 
             SenderBIC = 'MGTCBEBEXXX', 
             ana_dizin="Q:\OprsICM\YURTDIŞI SPOT\EUROCLEAR\ECL 2024",
             manuel_tarih_baslangici = manuel_tarih_baslangici ,
             manuel_tarih_bitisi     = manuel_tarih_bitisi,
             parcali_islemleri_birlestir= True)
      


# Clear Stream için Mutabakat yap
MutabakatYap(dosya_on_eki = "CBL",
             SenderBIC = 'CEDELULLXXX',
             ana_dizin="Q:\OprsICM\YURTDIŞI SPOT\CLEARSTREAM\CBL 2024", 
             manuel_tarih_baslangici = manuel_tarih_baslangici ,
             manuel_tarih_bitisi     = manuel_tarih_bitisi,
             parcali_islemleri_birlestir = False)






